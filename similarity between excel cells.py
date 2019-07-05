#!/usr/bin/python3
import openpyxl
import re

mainPath='/Users/diosim/Downloads/main.xlsx'
mainWb=openpyxl.load_workbook(mainPath)
mainSheet=mainWb['Sheet1']

superPath='/Users/diosim/Downloads/superset1.xlsx'
superWb=openpyxl.load_workbook(superPath)
superSheet=superWb['Sheet1']



#compare how similar the two strings are, return #.##
def similarity(s1, s2):
    s1 = (re.sub('\W+','', s1)).lower()
    s2 = (re.sub('\W+','', s2)).lower()
    return 2. * len(longest_common_substring(s1, s2)) / (len(s1) + len(s2)) * 100


def longest_common_substring(s1, s2):
   m = [[0] * (1 + len(s2)) for i in range(1 + len(s1))]
   longest, x_longest = 0, 0
   for x in range(1, 1 + len(s1)):
       for y in range(1, 1 + len(s2)):
           if s1[x - 1] == s2[y - 1]:
               m[x][y] = m[x - 1][y - 1] + 1
               if m[x][y] > longest:
                   longest = m[x][y]
                   x_longest = x
           else:
               m[x][y] = 0
   return s1[x_longest - longest: x_longest]



#Dataset contains 135869~ rows, smaller one to compare to has 2009~
for y in range (2,135869,1):
    print('current run: ', y)
    for z in range (2,2009,1):
        mainName = re.sub('\W+','', mainSheet.cell(row=z,column=2).value.upper())
        superName = re.sub('\W+','', superSheet.cell(row=y,column=15).value)
        if (mainSheet.cell(row=z,column=2).value.upper() == superSheet.cell(row=y,column=15).value) & int((mainSheet.cell(row=z,column=9).value) == int(superSheet.cell(row=y,column=19).value)):
            print('found a match!')
            #write in the smaller one r2c11 the zipcode from the dataset ryc31
            mainSheet.cell(row=z,column=11).value = superSheet.cell(row=y,column=31).value


#TEST AREA
#str1 = 'CABRINI UNIVERSITY'
#str2 = 'Cabrini College'
#print(similarity(str1, str2))


mainWb.save("filtered.xlsx")
print('job finished')
